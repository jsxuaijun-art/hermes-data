#!/usr/bin/env python3
"""
Convert 小企业会计准则 balance sheet (.xls) to 民间非营利组织 format (.xlsx).

Usage:
    python scripts/convert_minsheng_balance_sheet.py input.xls output.xlsx year

Dependencies: xlrd, openpyxl

Example:
    python scripts/convert_minsheng_balance_sheet.py \\
        /mnt/c/Users/jiangmin/Desktop/资产负债表2024.xls \\
        /mnt/c/Users/jiangmin/Desktop/资产负债表2024（民非）.xlsx 2024
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ── Styles ─────────────────────────────────────────────
THIN = Side(style='thin')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
FONT_TITLE = Font(name='微软雅黑', size=14, bold=True)
FONT_SUB = Font(name='微软雅黑', size=10)
FONT_HDR = Font(name='微软雅黑', size=9, bold=True, color='FFFFFF')
FONT_NORM = Font(name='微软雅黑', size=9)
FONT_BOLD = Font(name='微软雅黑', size=9, bold=True)
FONT_SML = Font(name='微软雅黑', size=8)
FILL_HDR = PatternFill('solid', fgColor='336699')
FILL_SEC = PatternFill('solid', fgColor='4472C4')
FILL_ALT = PatternFill('solid', fgColor='F2F2F2')
ALIGN_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_R = Alignment(horizontal='right', vertical='center', wrap_text=True)

def fmt(v):
    if v is None or abs(v) < 0.001:
        return None
    return round(v, 2)

def read_xls(path):
    import xlrd
    wb = xlrd.open_workbook(path)
    sh = wb.sheet_by_index(0)
    dv = sh.cell_value(2, 3)
    if isinstance(dv, float):
        t = xlrd.xldate_as_tuple(dv, wb.datemode)
        ds = f'{t[0]}年{t[1]}月{t[2]}日'
    else:
        ds = str(dv)
    rows = [tuple(sh.cell_value(r, c) for c in range(sh.ncols))
            for r in range(sh.nrows)]
    return rows, ds

def extract_raw(rows):
    g = lambda r, c: rows[r][c] if r < len(rows) else 0
    return {
        'money': (g(5,2), g(5,3)),
        'short_inv': (g(6,2), g(6,3)),
        'notes_rcv': (g(7,2), g(7,3)),
        'ar': (g(8,2), g(8,3)),
        'prepay': (g(9,2), g(9,3)),
        'other_rcv': (g(12,2), g(12,3)),
        'inventory': (g(13,2), g(13,3)),
        'other_ca': (g(18,2), g(18,3)),
        'ca_total': (g(19,2), g(19,3)),
        'lt_prepay': (g(32,2), g(32,3)),
        'fa_orig': (g(23,2), g(23,3)),
        'fa_depr': (g(24,2), g(24,3)),
        'fa_net': (g(25,2), g(25,3)),
        'nca_total': (g(34,2), g(34,3)),
        'asset_total': (g(35,2), g(35,3)),
        'short_loan': (g(5,6), g(5,7)),
        'notes_pay': (g(6,6), g(6,7)),
        'ap': (g(7,6), g(7,7)),
        'advances': (g(8,6), g(8,7)),
        'wages': (g(9,6), g(9,7)),
        'tax': (g(10,6), g(10,7)),
        'payable': (g(13,6), g(13,7)),
        'other_cl': (g(18,6), g(18,7)),
        'cl_total': (g(19,6), g(19,7)),
        'lt_loan': (g(21,6), g(21,7)),
        'lt_payable': (g(22,6), g(22,7)),
        'liab_total': (g(26,6), g(26,7)),
        'capital': (g(30,6), g(30,7)),
        'cap_reserve': (g(31,6), g(31,7)),
        'surplus': (g(32,6), g(32,7)),
        'retained': (g(33,6), g(33,7)),
        'equity_total': (g(34,6), g(34,7)),
    }

def get(d, k):
    return fmt(d[k][0]), fmt(d[k][1])

def build_minsheng(d):
    rcv_q, rcv_n = fmt(d['ar'][0]+d['prepay'][0]+d['other_rcv'][0]), fmt(d['ar'][1]+d['prepay'][1]+d['other_rcv'][1])
    pay_q, pay_n = fmt(d['ap'][0]+d['payable'][0]), fmt(d['ap'][1]+d['payable'][1])
    eq_q, eq_n = get(d, 'equity_total')
    a_q, a_n = get(d, 'asset_total')
    l_q, l_n = get(d, 'liab_total')

    left = [
        ('sec', '流动资产：'),
        ('it', '货币资金', 1, *get(d, 'money')),
        ('it', '短期投资', 2, None, None),
        ('it', '应收款项', 3, rcv_q, rcv_n),
        ('it', '预付账款', 4, None, None),
        ('it', '存货', 5, None, None),
        ('it', '待摊费用', 6, None, None),
        ('it', '一年内到期的长期债权投资', 7, None, None),
        ('it', '其他流动资产', 8, *get(d, 'other_ca')),
        ('ttl', '流动资产合计', 9, *get(d, 'ca_total')),
        ('bl',),
        ('sec', '非流动资产：'),
        ('it', '长期股权投资', 10, None, None),
        ('it', '长期债权投资', 11, None, None),
        ('dt', '固定资产原价', '', *get(d, 'fa_orig')),
        ('dt', '减：累计折旧', '', *get(d, 'fa_depr')),
        ('it', '固定资产净值', 12, *get(d, 'fa_net')),
        ('it', '无形资产', 13, None, None),
        ('it', '受托代理资产', 14, None, None),
        ('dt', '长期待摊费用', '', *get(d, 'lt_prepay')),
        ('it', '其他非流动资产', 15, None, None),
        ('ttl', '非流动资产合计', 16, *get(d, 'nca_total')),
        ('bl',),
        ('ttl', '资产总计', '', a_q, a_n),
    ]
    right = [
        ('sec', '流动负债：'),
        ('it', '短期借款', 17, *get(d, 'short_loan')),
        ('it', '应付款项', 18, pay_q, pay_n),
        ('it', '应付工资', 19, *get(d, 'wages')),
        ('it', '应交税金', 20, *get(d, 'tax')),
        ('it', '预收账款', 21, None, None),
        ('it', '预提费用', 22, None, None),
        ('it', '预计负债', 23, None, None),
        ('it', '一年内到期的长期负债', 24, None, None),
        ('it', '其他流动负债', 25, *get(d, 'other_cl')),
        ('ttl', '流动负债合计', 26, *get(d, 'cl_total')),
        ('sec', '长期负债：'),
        ('it', '长期借款', 27, *get(d, 'lt_loan')),
        ('it', '长期应付款', 28, *get(d, 'lt_payable')),
        ('it', '其他长期负债', 29, None, None),
        ('ttl', '长期负债合计', 30, None, None),
        ('bl',),
        ('it', '受托代理负债', '', None, None),
        ('ttl', '负债合计', '', l_q, l_n),
        ('bl',),
        ('sec', '净资产：'),
        ('it', '非限定性净资产', 31, eq_q, eq_n),
        ('it', '限定性净资产', 32, None, None),
        ('ttl', '净资产合计', 33, eq_q, eq_n),
        ('bl',),
        ('ttl', '负债和净资产总计', '', a_q, a_n),
    ]
    return left, right

def write_xlsx(left, right, date_str, out):
    wb = Workbook()
    ws = wb.active
    ws.title = '资产负债表'

    for c, w in {1:22,2:6,3:16,4:16,5:3,6:22,7:6,8:16,9:16}.items():
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 20

    def merge_set(c1, c2, val, font, align):
        ws.merge_cells(f'{c1}:{c2}')
        c = ws[c1]; c.value = val; c.font = font; c.alignment = align

    merge_set('A1', 'D1', '资产负债表', FONT_TITLE, ALIGN_C)
    merge_set('F1', 'I1', '资产负债表', FONT_TITLE, ALIGN_C)
    merge_set('A2', 'D2', '（适用民间非营利组织）', FONT_SUB, ALIGN_C)
    merge_set('F2', 'I2', '（适用民间非营利组织）', FONT_SUB, ALIGN_C)

    ws.merge_cells('A3:D3')
    ws['A3'] = '编制单位：苏州市姑苏区爱心之家老年公寓'
    ws['A3'].font = FONT_NORM; ws['A3'].alignment = ALIGN_L
    ws.merge_cells('F3:H3')
    ws['F3'] = date_str; ws['F3'].font = FONT_NORM; ws['F3'].alignment = ALIGN_C
    ws['I3'] = '单位：元'; ws['I3'].font = FONT_NORM; ws['I3'].alignment = ALIGN_R

    def set_cell(r, c, val=None, font=FONT_NORM, fill=None, align=ALIGN_C):
        cell = ws.cell(r, c)
        if val is not None: cell.value = val
        cell.font = font; cell.alignment = align; cell.border = BORDER
        if fill: cell.fill = fill

    for i, h in enumerate(['资 产', '行次', '期末余额', '年初余额']):
        set_cell(4, i+1, h, FONT_HDR, FILL_HDR)
    set_cell(4, 5, fill=FILL_HDR)
    for i, h in enumerate(['负债和净资产', '行次', '期末余额', '年初余额']):
        set_cell(4, i+6, h, FONT_HDR, FILL_HDR)

    n = max(len(left), len(right))
    while len(left) < n: left.append(('bl',))
    while len(right) < n: right.append(('bl',))

    for i in range(n):
        r = 5 + i
        ws.row_dimensions[r].height = 20
        fl = FILL_ALT if i % 2 else None
        lr, rr = left[i], right[i]

        # Left
        lt = lr[0]
        if lt == 'sec':
            ws.merge_cells(f'A{r}:D{r}')
            set_cell(r, 1, f'  {lr[1]}',
                     Font(name='微软雅黑', size=9, bold=True, color='FFFFFF'), FILL_SEC, ALIGN_L)
            for cc in [2,3,4]: set_cell(r, cc, fill=FILL_SEC)
        elif lt == 'bl':
            for cc in [1,2,3,4]: set_cell(r, cc)
        elif lt == 'dt':
            _, nm, rn, q, nc = lr
            set_cell(r, 1, f'    {nm}', FONT_SML, fl, ALIGN_L)
            set_cell(r, 2, rn, fill=fl)
            set_cell(r, 3, q, fill=fl, align=ALIGN_R)
            set_cell(r, 4, nc, fill=fl, align=ALIGN_R)
            ws.cell(r,3).number_format = '#,##0.00'
            ws.cell(r,4).number_format = '#,##0.00'
        else:
            is_ttl = lt == 'ttl'
            _, nm, rn, q, nc = lr
            f = FONT_BOLD if is_ttl else FONT_NORM
            set_cell(r, 1, f'  {nm}', f, fl, ALIGN_L)
            set_cell(r, 2, rn if rn else None, fill=fl)
            set_cell(r, 3, q, f if q is not None else FONT_NORM, fl, ALIGN_R)
            set_cell(r, 4, nc, f if nc is not None else FONT_NORM, fl, ALIGN_R)
            ws.cell(r,3).number_format = '#,##0.00'
            ws.cell(r,4).number_format = '#,##0.00'

        set_cell(r, 5, fill=fl)

        # Right
        rt = rr[0]
        if rt == 'sec':
            ws.merge_cells(f'F{r}:I{r}')
            set_cell(r, 6, f'  {rr[1]}',
                     Font(name='微软雅黑', size=9, bold=True, color='FFFFFF'), FILL_SEC, ALIGN_L)
            for cc in [7,8,9]: set_cell(r, cc, fill=FILL_SEC)
        elif rt == 'bl':
            for cc in [6,7,8,9]: set_cell(r, cc, fill=fl)
        else:
            is_ttl = rt == 'ttl'
            _, nm, rn, q, nc = rr
            f = FONT_BOLD if is_ttl else FONT_NORM
            set_cell(r, 6, f'  {nm}', f, fl, ALIGN_L)
            set_cell(r, 7, rn if rn else None, fill=fl)
            set_cell(r, 8, q, f if q is not None else FONT_NORM, fl, ALIGN_R)
            set_cell(r, 9, nc, f if nc is not None else FONT_NORM, fl, ALIGN_R)
            ws.cell(r,8).number_format = '#,##0.00'
            ws.cell(r,9).number_format = '#,##0.00'

    # Footer
    fr = 5 + n + 1
    ws.row_dimensions[fr].height = 25
    for cc in range(1, 10):
        ws.cell(fr, cc).border = Border(top=THIN)
    set_cell(fr, 1, '单位负责人：', align=ALIGN_L)
    ws.merge_cells(f'D{fr}:F{fr}')
    set_cell(fr, 4, '财务负责人：', align=ALIGN_C)
    ws.merge_cells(f'G{fr}:I{fr}')
    set_cell(fr, 7, '制表人：', align=ALIGN_R)

    ws.print_title_rows = '1:4'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    wb.save(out)
    return out

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print('Usage: python convert_minsheng_balance_sheet.py input.xls output.xlsx [year]')
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]
    year = int(sys.argv[3]) if len(sys.argv) > 3 else 0
    rows, ds = read_xls(src)
    if year: ds = f'{year}年12月31日'
    d = extract_raw(rows)
    left, right = build_minsheng(d)
    write_xlsx(left, right, ds, dst)
    print(f'✔ {dst}')
